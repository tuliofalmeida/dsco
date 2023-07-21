import os
import math
import random
import string
import openpyxl
import xlsxwriter
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.signal import find_peaks,peak_widths


def create_xlsx_citron(folder_path):
    """
    Func to create the excel file specific
    for Agathe data.
    
    Parameters
    --------------
    folder_path: str
        experiments folder, retrivied by the GUI
        
    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    xlsx_file = os.path.join(folder_path,folder_path.split(os.sep)[-1] + '_results.xlsx')
    if os.path.isfile(xlsx_file):
        pass
    else:
        workbook = xlsxwriter.Workbook(xlsx_file)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'File')
        worksheet.write(0, 1, 'Rostrocaudal Speed')
        worksheet.write(0, 2, 'Mediolateral Speed')
        worksheet.write(0, 3, 'Half Rise ROI 1')
        worksheet.write(0, 4, 'Half Rise ROI 2')
        worksheet.write(0, 5, 'Half Rise ROI 3')
        worksheet.write(0, 6, 'Half Rise ROI 4')
        worksheet.write(0, 7, 'Half Width ROI 1')
        worksheet.write(0, 8, 'Half Width ROI 2')
        worksheet.write(0, 9, 'Half Width ROI 3')
        worksheet.write(0, 10, 'Half Width ROI 4')
        worksheet.write(0, 11, 'Rise ROI 1')
        worksheet.write(0, 12, 'Rise ROI 2')
        worksheet.write(0, 13, 'Rise ROI 3')
        worksheet.write(0, 14, 'Rise ROI 4')
        worksheet.write(0, 15, 'Decay ROI 1')
        worksheet.write(0, 16, 'Decay ROI 2')
        worksheet.write(0, 17, 'Decay ROI 3')
        worksheet.write(0, 18, 'Decay ROI 4')
        worksheet.write(0, 19, 'Distance Midline')
        worksheet.write(0, 20, 'Rostrocaudal Distance')
        worksheet.write(0, 21, 'Mediolateral Distance')
        worksheet.write(0, 22, 'Frame used')
        worksheet.write(0, 23, 'Depolarization Start')
        workbook.close()

    return xlsx_file

def save_xlsx_citron(xlsx_file,idx,rostrocaudal_speed,mediolateral_speed,metrics,file_name,
                    rostrocaudal,mediolateral,frame_idx,origin,n_rois,coords_plot_w,px_value):
    """
    Func to save the excel file specific
    for Agathe data.
    
    Parameters
    --------------
    xlsx_file : str
        Path of the excel file
    idx : int
        Number of the file that is running (row)
    rostrocaudal_speed : float
        output of 'propagation_speed_4rois()'
    mediolateral_speed : float
        output of 'propagation_speed_4rois()'
    metrics : dict
        output of 'half_peak()'
    file_name : str
        name of the file to save
    rostrocaudal : float
        distance metric, output of 'distance_4rois()'
    mediolateral: float
        distance metric, output of 'distance_4rois()'
    frame_idx : int
        number of the frame used
    origin : str
        Where the wave starts
    n_rois : int
        Number of ROIs used
    coords_plot_w : float
        Width of the hidden ROI
    px_value :  float
        value to convert ixel to um

    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """   
    
    col = idx + 2

    xfile = openpyxl.load_workbook(xlsx_file)

    sheet = xfile['Sheet1']
    sheet['A{}'.format(col)] = file_name[idx]
    sheet['B{}'.format(col)] = rostrocaudal_speed
    sheet['C{}'.format(col)] = mediolateral_speed
    sheet['D{}'.format(col)] = metrics[1][0]
    sheet['E{}'.format(col)] = metrics[2][0]
    sheet['F{}'.format(col)] = metrics[3][0]
    sheet['G{}'.format(col)] = metrics[4][0]
    sheet['H{}'.format(col)] = metrics[1][1]
    sheet['I{}'.format(col)] = metrics[2][1]
    sheet['J{}'.format(col)] = metrics[3][1]
    sheet['K{}'.format(col)] = metrics[4][1]
    sheet['L{}'.format(col)] = metrics[1][2]
    sheet['M{}'.format(col)] = metrics[2][2]
    sheet['N{}'.format(col)] = metrics[3][2]
    sheet['O{}'.format(col)] = metrics[4][2]
    sheet['P{}'.format(col)] = metrics[1][3]
    sheet['Q{}'.format(col)] = metrics[2][3]
    sheet['R{}'.format(col)] = metrics[3][3]
    sheet['S{}'.format(col)] = metrics[4][3]
    try:
        sheet['T{}'.format(col)] = coords_plot_w[n_rois][2]*px_value
    except:
        sheet['T{}'.format(col)] = 'NO FIFTH ROI'
    try:
        sheet['U{}'.format(col)] = rostrocaudal
    except:
        sheet['U{}'.format(col)] = 'ROI ERROR'
    try:
        sheet['V{}'.format(col)] = mediolateral
    except:
        sheet['V{}'.format(col)] = 'ROI ERROR'
    sheet['W{}'.format(col)] = frame_idx
    sheet['X{}'.format(col)] = origin

    xfile.save(xlsx_file)

def create_xlsx_3ROI(folder_path):
    """
    Func to create the excel file for
    the 3 ROIs pipeline
    
    Parameters
    --------------
    folder_path : str
        experiments folder, retrivied by the GUI
        
    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    xlsx_file = os.path.join(folder_path,folder_path.split(os.sep)[-1] + '_3ROI_results.xlsx')
    if os.path.isfile(xlsx_file):
        pass
    else:
        workbook = xlsxwriter.Workbook(xlsx_file)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'File')
        worksheet.write(0, 1, 'Half Rise ROI 1')
        worksheet.write(0, 2, 'Half Rise ROI 2')
        worksheet.write(0, 3, 'Half Rise ROI 3')
        worksheet.write(0, 4, 'Half Width ROI 1')
        worksheet.write(0, 5, 'Half Width ROI 2')
        worksheet.write(0, 6, 'Half Width ROI 3')
        worksheet.write(0, 7, 'Rise ROI 1')
        worksheet.write(0, 8, 'Rise ROI 2')
        worksheet.write(0, 9, 'Rise ROI 3')
        worksheet.write(0, 10, 'Decay ROI 1')
        worksheet.write(0, 11, 'Decay ROI 2')
        worksheet.write(0, 12, 'Decay ROI 3')
        worksheet.write(0, 13, 'Frame used')
        workbook.close()

    return xlsx_file

def save_xlsx_3ROI(xlsx_file,idx,metrics,file_name,frame_idx):
    """
    Func to save the excel file specific
    for Agathe data.
    
    Parameters
    --------------
    xlsx_file : str
        Path of the excel file
    idx : int
        Number of the file that is running (row)
    metrics : dict
        output of 'half_peak()'
    file_name : str
        name of the file to save
    frame_idx : int
        number of the frame used

    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """   
    
    col = idx + 2

    xfile = openpyxl.load_workbook(xlsx_file)

    sheet = xfile['Sheet1']
    sheet['A{}'.format(col)] = file_name[idx]
    sheet['B{}'.format(col)] = metrics[1][0]
    sheet['C{}'.format(col)] = metrics[2][0]
    sheet['D{}'.format(col)] = metrics[3][0]
    sheet['E{}'.format(col)] = metrics[1][1]
    sheet['F{}'.format(col)] = metrics[2][1]
    sheet['G{}'.format(col)] = metrics[3][1]
    sheet['H{}'.format(col)] = metrics[1][2]
    sheet['I{}'.format(col)] = metrics[2][2]
    sheet['J{}'.format(col)] = metrics[3][2]
    sheet['K{}'.format(col)] = metrics[1][3]
    sheet['L{}'.format(col)] = metrics[2][3]
    sheet['M{}'.format(col)] = metrics[3][3]
    sheet['N{}'.format(col)] = frame_idx

    xfile.save(xlsx_file)

def create_xlsx(folder_path,n_rois):
    """
    Func to create the excel file for
    the general pipeline
    
    Parameters
    --------------
    folder_path : str
        experiments folder, retrivied by the GUI
    n_rois : int
        amount of ROIs used
        
    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    xlsx_file = os.path.join(folder_path,folder_path.split(os.sep)[-1] + '_results.xlsx')
    if os.path.isfile(xlsx_file):
        pass
    else:
        ref = np.asarray([3,4,5,6])
        workbook = xlsxwriter.Workbook(xlsx_file)
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'File')
        worksheet.write(0, 1, 'Distance Midline')
        worksheet.write(0, 2, 'Frame used')
        for roi in range(n_rois):
            worksheet.write(0, ref[0], 'Half Rise ROI {}'.format(roi+1))
            worksheet.write(0, ref[1], 'Half Width ROI {}'.format(roi+1))
            worksheet.write(0, ref[2], 'Rise ROI {}'.format(roi+1))
            worksheet.write(0, ref[3], 'Decay ROI {}'.format(roi+1))
            ref += 4
        ref = ref[:2]
        for hroi in range(0,n_rois,2):
            worksheet.write(0, ref[0], 'Distance ROI {}-{}'.format(hroi+1,hroi+2))
            worksheet.write(0, ref[1], 'Speed ROI {}-{}'.format(hroi+1,hroi+2))
            ref += 2
        workbook.close()
    
    return xlsx_file

def save_xlsx(xlsx_file,idx,metrics,dis_metrics,file_name,frame_idx,n_rois,coords_plot_w,px_value):    
    """
    Func to save the excel file specific
    for Agathe data.
    
    Parameters
    --------------
    xlsx_file : str
        Path of the excel file
    idx : int
        Number of the file that is running (row)
    metrics : dict
        output of 'half_peak()'
    dis_metrics : dict
        output of 'measurements()'
    file_name : str
        name of the file to save
    frame_idx : int
        number of the frame used
    n_rois : int
        Number of ROIs used
    coords_plot_w : float
        Width of the hidden ROI
    px_value :  float
        value to convert ixel to um

    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    xfile = openpyxl.load_workbook(xlsx_file)
    col = idx + 2
    ref = np.asarray([3,4,5,6])
    rows1 = {idx:letter for idx,letter in enumerate(string.ascii_uppercase)}
    rows2 = {idx+len(string.ascii_uppercase):'A'+letter for idx,letter in enumerate(string.ascii_uppercase)}
    rows3 = {idx+len(string.ascii_uppercase)*2:'B'+letter for idx,letter in enumerate(string.ascii_uppercase)}
    rows = {**rows1,**rows2,**rows3}
    sheet = xfile['Sheet1']
    sheet['A{}'.format(col)] = file_name[idx]
    try:
        sheet['B{}'.format(col)] = coords_plot_w[n_rois][2]*px_value
    except:
        sheet['B{}'.format(col)] = 'NO EXTRA ROI'
    sheet['C{}'.format(col)] = frame_idx
    for roi in range(n_rois):
        sheet['{}{}'.format(rows[ref[0]],col)] = metrics[roi+1][0]
        sheet['{}{}'.format(rows[ref[1]],col)] = metrics[roi+1][1]
        sheet['{}{}'.format(rows[ref[2]],col)] = metrics[roi+1][2]
        sheet['{}{}'.format(rows[ref[3]],col)] = metrics[roi+1][3]
        ref += 4
    ref = ref[:2]
    for hroi in range(0,n_rois,2):
        key = str(hroi+1)+'-'+str(hroi+2) 
        sheet['{}{}'.format(rows[ref[0]],col)] = dis_metrics[key][0]
        sheet['{}{}'.format(rows[ref[1]],col)] = dis_metrics[key][1]
        ref += 2

    xfile.save(xlsx_file)

def frame_plot(z_array):
    """
    This function find the frame with the highest
    pixel value. Returns the minimal pixel value,
    the highest pixel value and the frame index.
    
    Parameters
    --------------
    z_array: DataArray
        The images after all the processing. This
        array is converted before from Java to 
        Python.
        
    Returns
    --------------
    px_min : int
        value of the lowest pixel
    px_max : int
        value of highest pixel
    frame_idx : int
        index of the frame with px_max

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    px_min = 0
    maximas,means = [],[]
    for f in range(z_array.shape[0]):
        maximas.append(float(z_array[f].max()))
        means.append(np.mean(z_array[f]))   
        if float(z_array[f].min()) < px_min:
            px_min = float(z_array[f].min())

    max_peak, _ = find_peaks(means, width=3)
    px_max = max(np.asarray(maximas)[max_peak])
    frame_idx = np.where(np.asarray(maximas)==px_max)[0][0]
    
    return int(px_min), int(px_max), frame_idx

def rect_values(coord_right,coord_left):
    """
    This function creates the rectangle ROI based on 
    right mouse coordinate (coord_right) and left mouse 
    coordinate (coord_left). These lists comes from 
    onclick function. 
    
    Parameters
    --------------
    coord_right: list
        list with mouse right click coordinates (x,y)
    coord_left: list
        list with mouse left click coordinates (x,y)
        
    Returns
    --------------
    x : int
        value of down left x
    y : int
        value of down left y
    w : int
        width of the ROI
    h : int
        width of the ROI

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    x = min((coord_right[0],coord_left[0]))
    y = min((coord_right[1],coord_left[1]))
    xw = max((coord_right[0],coord_left[0]))
    yh = max((coord_right[1],coord_left[1]))
    w = xw-x
    h = yh-y

    return x,y,w,h

def determined_rois(coord_left,coord_right,n_rois):
    """
    This function creates a dictionary for all ROIs
    determined by the user containing the ROI number
    as a key (Python style) and as value a list with
    x and y coordinates for the upper left point of 
    the rectangle and the widht and height value in 
    pixels. This is done using the ImageJ logic to
    create ROIs.
    
    Parameters
    --------------
    coord_left : list
        list with the coordinates determined
        by the left mouse click
    coord_right : list
        list with the coordinates determined
        by the right mouse click
    n_rois : int
        the amount of ROIs used in the analysis
        
    Returns
    --------------
    roi_areas : dict
        Dict with list of rectangle info for
        each ROI

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    roi_areas = {}
    for roi in range(n_rois):
        x,y,w,h = rect_values(coord_right[roi],coord_left[roi])
        roi_areas[roi] = [x,y,w,h]
    
    return roi_areas

def plot_rois(roi_areas, frame, px_max, file_name, path2save):
    """
    this function will plot all rois determined
    by the user and will save it in the 'path2save'
    folder.
    
    Parameters
    --------------
    roi_areas : dict
        dict with the infos of each ROI, output of
        determined_rois
    frame : DataArray
        Matrix of the pixel with the highest value. 
        This variable is created in the code to 
        determine the ROIs, and its the 
        'z_array[frame_idx]'
    file_name : str
        Name of the file under analysis, variable
        created in the beggining of the code
    px_max : int
        the value of the highest pixel to normalize
        the plot
    path2save : str
        path of the analysis folder
        
    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    from matplotlib.patches import Rectangle

    fig,ax = plt.subplots()
    ax.imshow(frame,cmap = plt.cm.plasma, vmax = px_max)
    for roi in range(len(roi_areas)):
        ax.add_patch( Rectangle((roi_areas[roi][0], roi_areas[roi][1]),
                                roi_areas[roi][2], roi_areas[roi][3], 
                                facecolor= 'none', edgecolor = 'white'))
        ax.set_title("{} ROIs".format(file_name), fontsize = 17)
    fig.savefig(path2save + '_roi.svg')
    plt.close()

def roi_macro(roi_areas):
    """
    Function to create and run the macro that will
    create the ROIs determined by the user
    in the ImageJ and perform the multimeasure.
    Also, the macro will save the ROIs using the
    ImageJ configuration and the .csv with the 
    multimeasure data in the experiment folder
    pre-determined in the args dictionary.

    Parameters
    --------------
    roi_areas : dict
        dict with list of rectangle info for
        each ROI
        
    Returns
    --------------
    macro : str
        string with all commands in the ImageJ language

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    v1 = """#@ String inputFile\n""" 
    v2 = """#@ String outputFile\n"""
    v3 = """#@ String windowImg\n""" 
    v4 = """#@ String roiFile\n"""
    
    macro_init = """
    open(inputFile + ".tif");
    selectWindow(windowImg + ".tif");
    run("ROI Manager...");
    """

    macro_add = """"""
    for roi in range(len(roi_areas)):
        macro_add = macro_add + """
            makeRectangle({x}, {y}, {w}, {h});
            roiManager("Add");
            """.format(x = roi_areas[roi][0],
                       y = roi_areas[roi][1],
                       h = roi_areas[roi][3],
                       w = roi_areas[roi][2])
    # roiManager("Select", newArray(0,1,2,3));
    macro_end = """
    roiManager("Deselect");
    roiManager("Multi Measure");
    saveAs("Results", outputFile + ".csv");
    roiManager("Save", roiFile + "_roi.zip");
    close("*");
    close("ROI Manager");
    close("Results");
    """

    macro = v1 + v2 + v3 + v4 + macro_init + macro_add + macro_end

    return macro

def measuraments(df,center,px_value):
    """
    Function to calculate the distabce and
    the propagation speed between pair of ROIs.
    The estimation is performed using the the 
    CENTER of subsequents ROI, ex: the first 
    with the second; the third with the fourth
    and so on.

    Parameters
    --------------
    df : DataFrame
        DataFrame with the ROIs data
    center: dict
        Dict with the XY coords of the center for each ROI
    px_value: int (optional)
        Constant to convert pixel value to micrometers 
        
    Returns
    --------------
    dis_metrics : dict
        dict with the distance and the speed
        between the pairs, using the number of 
        ROIs as key. First value is the distance
        and the second the speed.

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    dis_metrics = {}
    distance = {'{}-{}'.format(xy+1,xy+2):math.dist(center[xy],center[xy+1])*px_value for xy in range(0,len(center),2)}
    
    for idxs, dis in distance.items():
        first = df['Mean{}'.format(idxs[0])]
        _, properties_f = find_peaks(first, prominence=max(first)//3, width=3)
        diff_first = properties_f["left_ips"]/10

        second = df['Mean{}'.format(idxs[2])]
        _, properties_s = find_peaks(second, prominence=max(second)//3, width=3)
        diff_second = properties_s["left_ips"]/10    

        speed = np.round(dis/abs(diff_first-diff_second),2)[0]

        dis_metrics[idxs] = [np.round(dis,2),speed]
    
    return dis_metrics

def distance_4rois(roi_areas, px_value):
    """
    This function estimates the distance between
    the 4 ROIs previously determined by the user.
    Returning the distance between the upper and
    lower roi (rostrocaudal) , and between the 
    lateral ROIs (mediolateral). The constant 
    'px_value' can be changed accordingly to the 
    experimental setup.

    Parameters
    --------------
    roi_areas: dict
        Dict with list of rectangle info for
        each ROI
    px_value: int (optional)
        Constant to convert pixel value to micrometers 
        
    Returns
    --------------
    rostrocaudal : float
        rostrocaudal distance in micrometers
    mediolateral : float
        mediolateral distance in micrometers

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    roi_x,roi_y = [],[]
    for idx, data in roi_areas.items():
        roi_x.append(data[0])
        roi_y.append(data[1])

    upper_roi = [roi_y.index(min(roi_y)),min(roi_y)]
    lower_roi = [roi_y.index(min(roi_y)),min(roi_y)]
    lateral = []

    for idx,y in enumerate(roi_y):
        if upper_roi[1] > y:
            upper_roi[0] = idx
            upper_roi[1] = y
        elif lower_roi[1] < y:
            lower_roi[0] = idx
            lower_roi[1] = y

    index = [lower_roi[0],upper_roi[0]]

    for idx in range(len(roi_x)):
        if idx not in index:
            lateral.append(idx)

    rostrocaudal = np.round(abs(upper_roi[1] - lower_roi[1]) * px_value,2)
    mediolateral = np.round(abs(roi_x[lateral[0]] - roi_x[lateral[1]]) * px_value,2)

    return rostrocaudal,mediolateral

def propagation_speed_4rois(df,roi_areas,n_rois,px_value):
    """
    Function to calculate the propagation speed of 
    the wave for all ROIs.

    Parameters
    --------------
    df : DataFrame
        DataFrame with the ROIs data
    roi_areas: dict
        Dict with list of rectangle info for
        each ROI
    n_rois : int
        the amount of ROIs used in the analysis
    px_value: int (optional)
        Constant to convert pixel value to micrometers 
        
    Returns
    --------------
    rostrocaudal_speed : float
        rostrocaudal propagation speed in um/s
    rostrocaudal_speed : float
        mediolateral propagation speed in um/s

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    diff_dict = {}
    rostrocaudal,mediolateral = distance_4rois(roi_areas,px_value)

    for roi in range(1,n_rois+1):
        column = 'Mean{}'.format(roi)
        y = df[column]
        _, properties = find_peaks(y, prominence=max(y)//3, width=3)
        for key,value in properties.items():
            if len(value)>1:
                    properties[key] = value[0]   
        try:   
            diff_dict[roi] = properties["left_ips"][0]/10
        except:
            diff_dict[roi] = properties["left_ips"]/10

    rostrocaudal_speed = rostrocaudal/abs(diff_dict[1]-diff_dict[3])
    mediolateral_speed = mediolateral/abs(diff_dict[2]-diff_dict[4])

    return np.round(rostrocaudal_speed,2), np.round(mediolateral_speed,2)

def load_data(path, n_rois, fps = 10):
    """
    Function to load the data processed using 
    the ROI pipeline. This function is to load 
    the data WITHOUT bootstrapping.

    Parameters
    --------------
    path : str
        data csv path.
    n_rois : int 
        the number of ROIs used in the analysis
    fps: int (optional)
        frames per second rate used in the experiment, used
        to create the time vector. Default = 10. 
        
    Returns
    --------------
    df : pd.DataFrame
        DataFrame with the time column and the mean value of 
        each ROI for each frame.

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    columns_list = ['Mean'+str(roi) for roi in range(1,n_rois+1)]
    df = pd.read_csv(path, usecols = columns_list)
    df.insert(0, "Time", np.arange(0,len(df)) * 1/fps)

    return df

def get_width(x, y, height = 0.5):
    """
    This function estimates the half height of the
    peak, returning the two positions in X and
    the width of this points (in seconds). 
    
    Parameters
    --------------
    x : nd.array
        Time column from the loaded df
    y : nd.array
        Mean column from the loaded df
    height : float (optional)
        Height of the estimation, pre determined
        the half peak (0.5)
        
    Returns
    --------------
    low : int
        Index of the rise of the curve
    high : int
        Index of the decay of the curve
    width : float
        Width of the curve

    Authors
    ------------
    Agathe Lafont and Tulio Almeida, adapted
    from https://stackoverflow.com/questions/10582795/finding-the-full-width-half-maximum-of-a-peak        
    """
    height_half_max = np.max(y) * height
    index_max = np.argmax(y)
    x_low = np.interp(height_half_max, y[:index_max+1], x[:index_max+1])
    x_high = np.interp(height_half_max, np.flip(y[index_max:]), np.flip(x[index_max:]))
    width = x_high-x_low
    low = (np.abs(x - x_low)).argmin()
    high = (np.abs(x - x_high)).argmin()

    return low,high,width

def roi_macro_bootstrap(roi_areas,px = 2,trials = 500):
    """
    Function to create and run the macro that will
    create the ROIs determined by the user
    in the ImageJ and perform the bootstrap (accordingly
    to the 'trials' parameterthe) and perform
    the multimeasure for each bootstraped ROI.
    Also, the macro will save the ROIs using the
    ImageJ configuration and the .csv with the 
    multimeasure data in the experiment folder
    pre-determined in the args dictionary.

    Parameters
    --------------
    roi_areas : dict
        dict with list of rectangle info for each ROI
    px : int (optional)
        variability to be add in the original ROI determined
        by the user. Default = 2.
    trials : int (optional)
        Amount of ROIs that will be created. Default = 500. 
        
    Returns
    --------------
    macro_dict : dict
        dictionary with the ImageJ commands to do the 
        boostrap for each ROI stored in 'roi_areas'

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    macro_dict = {}
    rois_boot = {}
    v1 = """#@ String inputFile\n""" 
    v2 = """#@ String outputFile\n"""
    v3 = """#@ String windowImg\n""" 
    v4 = """#@ String roiFile\n"""
    
    macro_init = """
    open(inputFile + ".tif");
    selectWindow(windowImg + ".tif");
    run("ROI Manager...");
    """

    for roi in range(len(roi_areas)):
        macro_add = """"""
        fake_rois = fakestrap(roi_areas[roi],px = px, trials = trials)
        rois_boot[roi] = [np.mean([xy[0] for xy in fake_rois]),
                          np.mean([xy[1] for xy in fake_rois]),
                          0,0]
        for fake_roi in range(len(fake_rois)):
            macro_add = macro_add + """
                makeRectangle({x}, {y}, {w}, {h});
                roiManager("Add");
                """.format(x = fake_rois[fake_roi][0],
                        y = fake_rois[fake_roi][1],
                        h = fake_rois[fake_roi][3],
                        w = fake_rois[fake_roi][2])

        macro_end = """
            roiManager("Deselect");
            roiManager("Multi Measure");
            saveAs("Results", outputFile + "_roi{}.csv");
            roiManager("Save", roiFile + "_roi{}.zip");
            close("*");
            close("ROI Manager");
            close("Results");
            """.format(str(roi+1),str(roi+1))

        macro = v1 + v2 + v3 + v4 + macro_init + macro_add + macro_end
        macro_dict[roi] = macro

    return macro_dict,rois_boot

def fakestrap(roi_area,px = 2,trials = 500):
    """
    This function works similar to a boostrapfunction,
    here will be created N ROIs (trials) using the pixel
    variability determined (px). After this is possible to
    calculate the mean between all ROIs and achive a more
    accurate value.

    Parameters
    --------------
    roi_area : str
        data csv path for the specific bootstraped ROI.
    px : int (optional)
        variability to be add in the original ROI determined
        by the user. Default = 2.
    trials : int (optional)
        Amount of ROIs that will be created. Default = 500. 
        
    Returns
    --------------
    df : pd.DataFrame
        DataFrame with the time column and the mean value of 
        each ROI created by the boostrap function for one
        original ROI determined

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    random_picks = {}
    for idx,point in enumerate(roi_area):
        temp = []
        window = np.arange(point-px,point+(px+1))
        for times in range(trials):
            pick = random.choice(window)
            temp.append(pick)
        random_picks[idx] = temp

    return list(zip(random_picks[0],random_picks[1],random_picks[2],random_picks[3]))

def _load_data_fakestrap(path, trials, fps = 10):
    """
    Internal use function. This function is 
    used to load the data processed using 
    the ROI pipeline with bootstrap.

    Parameters
    --------------
    path : str
        data csv path for the specific bootstraped ROI.
    trials : int 
        the amount of trials passed in the bootstrap function
    fps : int (optional)
        frames per second rate used in the experiment, used
        to create the time vector. Default = 10. 
        
    Returns
    --------------
    df : pd.DataFrame
        DataFrame with the time column and the mean value of 
        each ROI created by the boostrap function for one
        original ROI determined

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    columns_list = ['Mean'+str(roi) for roi in range(1,trials+1)]
    df = pd.read_csv(path, usecols = columns_list)
    df.insert(0, "Time", np.arange(0,len(df)) * 1/fps)
    
    return df

def load_data_bootstrap(path,n_rois,trials,fps = 10):
    """
    This function is used to load the data processed using 
    the ROI pipeline with bootstrap. Each row of each ROI
    will be the mean from all ROIs created by the bootstrap
    function.

    Parameters
    --------------
    path : str
        path of the experiment folder where all the csv's are
        stored
    n_rois : int
        the number of ROIs used in the analysis
    trials : int 
        the amount of trials passed in the bootstrap function
    fps : int (optional)
        frames per second rate used in the experiment, used
        to create the time vector. Default = 10. 
        
    Returns
    --------------
    df : pd.DataFrame
        DataFrame with the time column and the mean value of 
        each ROI created by the boostrap function.

    Authors
    ------------
    Agathe Lafont and Tulio Almeida        
    """
    df = pd.DataFrame()
    for roi in range(n_rois):
        df_temp = _load_data_fakestrap(path=path+"_roi{}.csv".format(str(roi+1)),trials=trials,fps=fps)
        df['Mean{}'.format(str(roi+1))] = df_temp.mean(axis=1)
    df.insert(0, "Time", np.arange(0,len(df)) * 1/fps)

    return df  

def plot_peaks(df, n_rois,plot_title,fps = 10,ax = None): 
    """
    Function to plot the curve of each ROI with 
    the peak, half peak and the witdh of the half
    peak (time).
    
    Parameters
    --------------
    df : DataFrame
        DataFrame with the ROIs data
    n_rois : int
        the amount of ROIs used in the analysis
    plot_title : str
        the title of the plot
    fps: int (optional)
        frames per second rate used in the experiment, used
        to create the time vector. Default = 10. 
    ax : nd.array (optional)
        the ax position to plot

    Returns
    --------------
    None

    Authors
    ------------
    Agathe Lafont and Tulio Almeida  
    """
    time = np.arange(0,len(df.Time)+2,15) # check
    if ax is None:
        for roi in range(1,n_rois+1):
            column = 'Mean{}'.format(roi)
            y = df[column]
            peaks, properties = find_peaks(y, prominence=max(y)//3, width=1)
            pw = peak_widths(y, peaks, rel_height=.9)
            pt = peak_widths(y, peaks, rel_height=.1)
            base = peak_widths(y, peaks, rel_height=.98)
            for key,value in properties.items():
                if len(value)>1:
                        properties[key] = value[0]
            plt.plot(y,color = 'k')
            if len(peaks) == 0:
                pass
            elif len(peaks)>1:
                plt.plot(peaks[0], y[peaks[0]], "x", color = 'C{}'.format(roi))
                # plt.plot(pw[2][0],pw[1][0],'o',color = 'C{}'.format(roi))
            else:
                plt.plot(peaks, y[peaks], "x", color = 'C{}'.format(roi))
                # plt.plot(pw[2],pw[1],'o',color = 'C{}'.format(roi))
            try:
                diff = np.round(df['Time'][int(properties["right_ips"])]-df['Time'][int(properties["left_ips"])],2)
                plt.hlines(y=properties["width_heights"], xmin=properties["left_ips"],
                        xmax=properties["right_ips"], color = 'C{}'.format(roi), 
                        label = 'ROI {} width {}s'.format(roi,diff),linestyles='dotted')
                plt.hlines(*pw[1:], color = 'C{}'.format(roi))
                plt.hlines(*pt[1:], color = 'C{}'.format(roi))
                plt.plot(base[2], base[1], "x", color = 'C{}'.format(roi))
            except:
                pass
            plt.xlabel('Time (s)',fontsize = 15)
            plt.ylabel('Pixel value',fontsize = 15)
            plt.xticks(time,labels = time*(1/fps), rotation = 45)
            plt.legend()
    else:
        out = []
        for roi in range(1,n_rois+1):
            column = 'Mean{}'.format(roi)
            y = df[column]
            peaks, properties = find_peaks(y, prominence=max(y)//3, width=3)
            pw = peak_widths(y, peaks, rel_height=.9)
            pt = peak_widths(y, peaks, rel_height=.1)
            base = peak_widths(y, peaks, rel_height=.98)
            for key,value in properties.items():
                if len(value)>1:
                        properties[key] = value[0]
            out.append(ax.plot(y,color = 'k'))
            if len(peaks) == 0:
                pass
            elif len(peaks)>1:
                out.append(ax.plot(peaks[0], y[peaks[0]], "x", color = 'C{}'.format(roi)))
                # out.append(ax.plot(pw[2][0],pw[1][0],'o',color = 'C{}'.format(roi)))
            else:
                out.append(ax.plot(peaks, y[peaks], "x", color = 'C{}'.format(roi)))
                # out.append(ax.plot(pw[2],pw[1],'o',color = 'C{}'.format(roi)))
            try:
                diff = np.round(df['Time'][int(properties["right_ips"])]-df['Time'][int(properties["left_ips"])],2)
                out.append(ax.hlines(y=properties["width_heights"], xmin=properties["left_ips"],
                            xmax=properties["right_ips"], color = 'C{}'.format(roi), 
                            label = 'ROI {} width {}s'.format(roi,diff),linestyles='dotted'))
                out.append(ax.hlines(*pw[1:], color = 'C{}'.format(roi)))
                out.append(ax.hlines(*pt[1:], color = 'C{}'.format(roi)))
                out.append(ax.plot(base[2], base[1], "x", color = 'C{}'.format(roi)))
            except:
                pass
            out.append(ax.set_xlabel('Time (s)',fontsize = 15))
            out.append(ax.set_ylabel('Pixel value',fontsize = 15))
            out.append(ax.set_xticks(time,labels = time*(1/fps), rotation = 45))
            out.append(ax.set_title(plot_title,fontsize = 17))
            out.append(ax.legend())
            
        return out
    
def half_peak(df, n_rois):
    """
    Function to get the half peak of each ROI,
    and determine the origin of the depolarization.
    
    Parameters
    --------------
    df : DataFrame
        DataFrame with the ROIs data
    n_rois : int
        the amount of ROIs used in the analysis

    Returns
    --------------
    half_peak_list : list
        half peak values (half peak - plot elbow base)
        for each ROI
    origin : str
        Direction of the depolarization

    Authors
    ------------
    Agathe Lafont and Tulio Almeida  
    """
    metrics,origins = {},[]

    origin = 'Rostral'
    for roi in range(1,n_rois+1):
        metrics[roi] = []
        column = 'Mean{}'.format(roi)
        y = df[column]
        peaks, properties = find_peaks(y, prominence=max(y)//3, width=3)
        base = peak_widths(y, peaks, rel_height=.98)
        pw = peak_widths(y, peaks, rel_height=.9)
        pt = peak_widths(y, peaks, rel_height=.1)
        rise = (pt[2]-pw[2])/10
        decay = (pw[3]-pt[3])/10
        try: 
            diff = np.round(df['Time'][int(properties["right_ips"])]-df['Time'][int(properties["left_ips"])],2)
        except:
            diff = np.round(df['Time'][int(properties["right_ips"][0])]-df['Time'][int(properties["left_ips"][0])],2)
        try:
            metrics[roi].append((properties['left_ips'][0]-base[2][0])/10)
            origins.append(properties['left_ips'][0])
        except:
            metrics[roi].append((properties['left_ips']-base[2])/10)
            origins.append(properties['left_ips'])
        metrics[roi].append(diff)
        metrics[roi].append(rise[0])
        metrics[roi].append(decay[0])

    if origins[0]-origins[2] > 0:
        origin = 'Caudal'

    return metrics,origin